#!/usr/bin/env python3

from __future__ import annotations

import json
import math
import struct
import subprocess
import zlib
from pathlib import Path
from typing import Iterable

import numpy as np
import openpyxl


ROOT = Path(__file__).resolve().parents[1]
GRAIN_DIR = ROOT / "public" / "images" / "grain"
BASELINE_XLSX = ROOT / "data" / "baseline.xlsx"
FUTURE_XLSX = ROOT / "data" / "future.xlsx"
METADATA_JSON = GRAIN_DIR / "paddy-map-metadata.json"
PROVINCE_JSON = GRAIN_DIR / "vietnam-provinces-map.json"
ALPHA_SOURCE = GRAIN_DIR / "paddy-baseline-2025-preview.png"

SCENARIOS = {
  "baseline": {
    "source": "baseline",
    "year": None,
    "preview": GRAIN_DIR / "paddy-baseline-2025-preview.png",
    "zoomPreview": GRAIN_DIR / "paddy-baseline-2025-preview-2x.png",
    "full": GRAIN_DIR / "paddy-baseline-2025.png",
  },
  "rcp45": {
    "source": "future",
    "year": 2050,
    "preview": GRAIN_DIR / "paddy-rcp45-2050-preview.png",
    "zoomPreview": GRAIN_DIR / "paddy-rcp45-2050-preview-2x.png",
    "full": GRAIN_DIR / "paddy-rcp45-2050.png",
  },
  "rcp85": {
    "source": "future",
    "year": 2050,
    "preview": GRAIN_DIR / "paddy-rcp85-2050-preview.png",
    "zoomPreview": GRAIN_DIR / "paddy-rcp85-2050-preview-2x.png",
    "full": GRAIN_DIR / "paddy-rcp85-2050.png",
  },
}

COLORS = {
  "green": (94, 169, 90),
  "yellow": (224, 194, 74),
  "red": (216, 83, 43),
}

SAMPLE_GRID_PIXELS = 96
INTERPOLATION_NEIGHBORS = 24
MAX_SEARCH_RING = 8

Sample = tuple[float, float, float]
SampleGrid = dict[tuple[int, int], list[Sample]]


def command(args: list[str], *, input_bytes: bytes | None = None) -> bytes:
  result = subprocess.run(args, input=input_bytes, stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=False)
  if result.returncode != 0:
    raise RuntimeError(f"{' '.join(args)}\n{result.stderr.decode('utf-8', errors='ignore')}")
  return result.stdout


def image_size(path: Path) -> tuple[int, int]:
  output = command(["magick", "identify", "-format", "%w %h", str(path)]).decode()
  width, height = output.split()
  return int(width), int(height)


def image_rgba(path: Path) -> tuple[int, int, bytes]:
  width, height = image_size(path)
  data = command(["magick", str(path), "-depth", "8", "rgba:-"])
  return width, height, data


def image_rgba_resized(path: Path, width: int, height: int) -> tuple[int, int, bytes]:
  data = command(["magick", str(path), "-resize", f"{width}x{height}!", "-depth", "8", "rgba:-"])
  return width, height, data


def write_png(path: Path, width: int, height: int, rgba: bytearray) -> None:
  def chunk(kind: bytes, data: bytes) -> bytes:
    return (
      struct.pack(">I", len(data))
      + kind
      + data
      + struct.pack(">I", zlib.crc32(kind + data) & 0xFFFFFFFF)
    )

  raw = bytearray()
  row_bytes = width * 4
  for y in range(height):
    raw.append(0)
    raw.extend(rgba[y * row_bytes : (y + 1) * row_bytes])

  png = (
    b"\x89PNG\r\n\x1a\n"
    + chunk(b"IHDR", struct.pack(">IIBBBBB", width, height, 8, 6, 0, 0, 0))
    + chunk(b"IDAT", zlib.compress(bytes(raw), 9))
    + chunk(b"IEND", b"")
  )
  path.write_bytes(png)


def mercator_y(latitude: float) -> float:
  radians = math.radians(max(-85.05112878, min(85.05112878, latitude)))
  return math.log(math.tan(math.pi / 4 + radians / 2))


with METADATA_JSON.open() as file:
  metadata = json.load(file)

BBOX = metadata["bbox"]
MERCATOR_TOP = mercator_y(BBOX["latMax"])
MERCATOR_BOTTOM = mercator_y(BBOX["latMin"])


def lon_lat_to_pixel(longitude: float, latitude: float, width: int, height: int) -> tuple[float, float]:
  x = ((longitude - BBOX["lonMin"]) / (BBOX["lonMax"] - BBOX["lonMin"])) * (width - 1)
  y = ((MERCATOR_TOP - mercator_y(latitude)) / (MERCATOR_TOP - MERCATOR_BOTTOM)) * (height - 1)
  return x, y


def load_baseline_rows() -> list[dict[str, float]]:
  workbook = openpyxl.load_workbook(BASELINE_XLSX, data_only=True, read_only=True)
  sheet = workbook.active
  rows = sheet.iter_rows(values_only=True)
  headers = [str(value) if value is not None else "" for value in next(rows)]
  index = {header: column for column, header in enumerate(headers)}
  parsed = []

  for values in rows:
    if values[index["Latitude"]] is None or values[index["Longitude"]] is None or values[index["Grain.As"]] is None:
      continue

    parsed.append(
      {
        "latitude": float(values[index["Latitude"]]),
        "longitude": float(values[index["Longitude"]]),
        "value": float(values[index["Grain.As"]]),
      }
    )

  return parsed


def load_future_rows(scenario: str, year: int) -> list[dict[str, float]]:
  workbook = openpyxl.load_workbook(FUTURE_XLSX, data_only=True, read_only=True)
  sheet = workbook.active
  rows = sheet.iter_rows(values_only=True)
  headers = [str(value) if value is not None else "" for value in next(rows)]
  index = {header: column for column, header in enumerate(headers)}
  parsed = []

  for values in rows:
    if values[index["scenario"]] != scenario or int(values[index["Year"]]) != year:
      continue

    parsed.append(
      {
        "latitude": float(values[index["Latitude"]]),
        "longitude": float(values[index["Longitude"]]),
        "value": float(values[index["Grain.As"]]),
      }
    )

  return parsed


def color_for_value(value: float) -> tuple[int, int, int]:
  if value <= 0.2:
    return COLORS["green"]
  if value <= 0.35:
    return COLORS["yellow"]
  return COLORS["red"]


def prepare_samples(rows: Iterable[dict[str, float]], width: int, height: int) -> list[Sample]:
  samples = []
  for row in rows:
    x, y = lon_lat_to_pixel(row["longitude"], row["latitude"], width, height)
    samples.append((x, y, row["value"]))
  return samples


def interpolate(x: float, y: float, samples: list[Sample]) -> float:
  weighted_sum = 0.0
  weight_total = 0.0

  for sample_x, sample_y, value in samples:
    distance_sq = (x - sample_x) ** 2 + (y - sample_y) ** 2
    if distance_sq < 0.0001:
      return value

    weight = 1.0 / distance_sq
    weighted_sum += weight * value
    weight_total += weight

  return weighted_sum / weight_total


def build_sample_grid(samples: list[Sample]) -> SampleGrid:
  sample_grid: SampleGrid = {}

  for sample in samples:
    sample_x, sample_y, _ = sample
    key = (int(sample_x // SAMPLE_GRID_PIXELS), int(sample_y // SAMPLE_GRID_PIXELS))
    sample_grid.setdefault(key, []).append(sample)

  return sample_grid


def nearby_samples(x: float, y: float, sample_grid: SampleGrid) -> list[Sample]:
  grid_x = int(x // SAMPLE_GRID_PIXELS)
  grid_y = int(y // SAMPLE_GRID_PIXELS)
  candidates: list[Sample] = []

  for ring in range(MAX_SEARCH_RING + 1):
    for cell_y in range(grid_y - ring, grid_y + ring + 1):
      for cell_x in range(grid_x - ring, grid_x + ring + 1):
        if ring > 0 and grid_x - ring < cell_x < grid_x + ring and grid_y - ring < cell_y < grid_y + ring:
          continue

        candidates.extend(sample_grid.get((cell_x, cell_y), []))

    if len(candidates) >= INTERPOLATION_NEIGHBORS:
      break

  if not candidates:
    for bucket in sample_grid.values():
      candidates.extend(bucket)

  candidates.sort(key=lambda sample: (x - sample[0]) ** 2 + (y - sample[1]) ** 2)
  return candidates[:INTERPOLATION_NEIGHBORS]


def interpolate_indexed(x: float, y: float, sample_grid: SampleGrid) -> float:
  weighted_sum = 0.0
  weight_total = 0.0

  for sample_x, sample_y, value in nearby_samples(x, y, sample_grid):
    distance_sq = (x - sample_x) ** 2 + (y - sample_y) ** 2
    if distance_sq < 0.0001:
      return value

    weight = 1.0 / distance_sq
    weighted_sum += weight * value
    weight_total += weight

  return weighted_sum / weight_total


def render_preview(path: Path, rows: list[dict[str, float]], alpha_rgba: bytes, width: int, height: int) -> None:
  samples = prepare_samples(rows, width, height)
  sample_x = np.asarray([sample[0] for sample in samples], dtype=np.float32)
  sample_y = np.asarray([sample[1] for sample in samples], dtype=np.float32)
  sample_values = np.asarray([sample[2] for sample in samples], dtype=np.float32)

  alpha = np.frombuffer(alpha_rgba, dtype=np.uint8).reshape((height, width, 4))[:, :, 3].reshape(-1)
  pixel_indexes = np.flatnonzero(alpha > 0)
  pixel_x = (pixel_indexes % width).astype(np.float32)
  pixel_y = (pixel_indexes // width).astype(np.float32)
  values = np.empty(pixel_indexes.shape[0], dtype=np.float32)

  chunk_size = 4096
  for start in range(0, pixel_indexes.shape[0], chunk_size):
    end = min(start + chunk_size, pixel_indexes.shape[0])
    dx = pixel_x[start:end, None] - sample_x[None, :]
    dy = pixel_y[start:end, None] - sample_y[None, :]
    distance_sq = dx * dx + dy * dy
    exact = distance_sq < 0.0001
    weights = 1.0 / np.maximum(distance_sq, 0.0001)
    chunk_values = (weights * sample_values[None, :]).sum(axis=1) / weights.sum(axis=1)

    if exact.any():
      exact_rows = exact.any(axis=1)
      chunk_values[exact_rows] = sample_values[exact[exact_rows].argmax(axis=1)]

    values[start:end] = chunk_values

  output = np.zeros((width * height, 4), dtype=np.uint8)
  output[pixel_indexes, 0] = np.where(values <= 0.2, COLORS["green"][0], np.where(values <= 0.35, COLORS["yellow"][0], COLORS["red"][0]))
  output[pixel_indexes, 1] = np.where(values <= 0.2, COLORS["green"][1], np.where(values <= 0.35, COLORS["yellow"][1], COLORS["red"][1]))
  output[pixel_indexes, 2] = np.where(values <= 0.2, COLORS["green"][2], np.where(values <= 0.35, COLORS["yellow"][2], COLORS["red"][2]))
  output[pixel_indexes, 3] = np.clip(alpha[pixel_indexes], 120, 212)

  write_png(path, width, height, bytearray(output.tobytes()))
  print(f"Wrote {path.relative_to(ROOT)} from {len(samples)} samples across {pixel_indexes.shape[0]} paddy pixels")


def update_full_from_preview(preview: Path, full: Path) -> None:
  width = metadata["rasterDimensions"]["width"]
  height = metadata["rasterDimensions"]["height"]
  command(["magick", str(preview), "-resize", f"{width}x{height}!", str(full)])
  print(f"Wrote {full.relative_to(ROOT)}")


def update_province_metrics(sample_sets: dict[str, list[dict[str, float]]], width: int, height: int) -> None:
  province_data = json.loads(PROVINCE_JSON.read_text())
  prepared = {scenario: prepare_samples(rows, width, height) for scenario, rows in sample_sets.items()}

  for province in province_data["provinces"]:
    center = province["center"]
    province["metrics"] = {
      scenario: round(interpolate(float(center["x"]), float(center["y"]), samples), 3)
      for scenario, samples in prepared.items()
    }

  PROVINCE_JSON.write_text(json.dumps(province_data, ensure_ascii=False, separators=(",", ":")))
  print(f"Updated {PROVINCE_JSON.relative_to(ROOT)} province metrics")


def main() -> None:
  width, height, alpha_rgba = image_rgba(ALPHA_SOURCE)
  zoom_width = width * 2
  zoom_height = height * 2
  _, _, zoom_alpha_rgba = image_rgba_resized(ALPHA_SOURCE, zoom_width, zoom_height)
  full_width = metadata["rasterDimensions"]["width"]
  full_height = metadata["rasterDimensions"]["height"]
  _, _, full_alpha_rgba = image_rgba_resized(ALPHA_SOURCE, full_width, full_height)
  sample_sets = {
    "baseline": load_baseline_rows(),
    "rcp45": load_future_rows("rcp45", 2050),
    "rcp85": load_future_rows("rcp85", 2050),
  }

  for scenario, config in SCENARIOS.items():
    render_preview(config["preview"], sample_sets[scenario], alpha_rgba, width, height)
    render_preview(config["zoomPreview"], sample_sets[scenario], zoom_alpha_rgba, zoom_width, zoom_height)
    render_preview(config["full"], sample_sets[scenario], full_alpha_rgba, full_width, full_height)

  update_province_metrics(sample_sets, width, height)


if __name__ == "__main__":
  main()
